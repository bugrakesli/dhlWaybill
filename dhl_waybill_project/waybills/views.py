import re
from decimal import Decimal, InvalidOperation
from datetime import date
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from django.db import transaction
from django.db.models import Sum, Count, Q
from django.http import HttpResponse
from django.utils.dateparse import parse_date
from rest_framework import status, generics, filters
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser

from .models import Waybill
from .serializers import WaybillSerializer, WaybillExcelUploadSerializer
from .pagination import WaybillPagination

PLACEHOLDER_DATE = date(1900, 1, 1)
PLACEHOLDER_TEXT = "-"


# --------------------------------------------------------------------------
# ORTAK YARDIMCI FONKSİYON: Filtreleme mantığı
# --------------------------------------------------------------------------

def filter_waybills_queryset(request):
    """
    start_date, end_date, delivered ve incomplete query parametrelerine göre
    Waybill queryset'ini filtreler.
    """
    queryset = Waybill.objects.all()

    start_date_param = request.query_params.get("start_date")
    end_date_param = request.query_params.get("end_date")
    delivered_param = request.query_params.get("delivered")
    incomplete_param = request.query_params.get("incomplete")

    start_date = parse_date(start_date_param) if start_date_param else None
    end_date = parse_date(end_date_param) if end_date_param else None

    if start_date:
        queryset = queryset.filter(shipment_date__gte=start_date)
    if end_date:
        queryset = queryset.filter(shipment_date__lte=end_date)

    if delivered_param is not None:
        delivered_str = str(delivered_param).strip().lower()
        if delivered_str in ["true", "1", "evet", "yes"]:
            queryset = queryset.filter(delivered=True)
        elif delivered_str in ["false", "0", "hayir", "hayır", "no"]:
            queryset = queryset.filter(delivered=False)

    # Eksik veri filtresi
    if incomplete_param and incomplete_param.lower() == "true":
        queryset = queryset.filter(
            Q(shipment_date=PLACEHOLDER_DATE)
            | Q(sender=PLACEHOLDER_TEXT)
            | Q(receiver=PLACEHOLDER_TEXT)
            | Q(destination=PLACEHOLDER_TEXT)
            | Q(collected_by=PLACEHOLDER_TEXT)
            | Q(euro_amount__isnull=True)
            | Q(exchange_rate__isnull=True)
            | Q(piece_count__isnull=True)
        )

    return queryset


# --------------------------------------------------------------------------
# 1) EXCEL YÜKLEME ENDPOINT'İ
# --------------------------------------------------------------------------

class WaybillExcelUploadView(APIView):
    """
    POST /api/waybills/upload/

    Excel/CSV (.xlsx, .xls, .csv) dosyasını okuyup Waybill kayıtlarını toplu olarak
    günceller (var olan waybill_number) veya oluşturur (yeni kayıt).
    """
    parser_classes = [MultiPartParser, FormParser]

    EMPTY_PLACEHOLDER_VALUES = {"", "NULL", "NAN", "N/A", "NA", "-", "NONE"}

    # Türkçe ve İngilizce sütun başlıklarını model alanlarına eşleme
    COLUMN_ALIASES = {
        # waybill_number (ZORUNLU)
        "AWB": "waybill_number",
        "KONŞİMENTO NO": "waybill_number",
        "KONSIMENTO NO": "waybill_number",
        "KONŞİMENTO": "waybill_number",
        "KONSIMENTO": "waybill_number",
        "WAYBILL_NUMBER": "waybill_number",
        "WAYBILL NO": "waybill_number",
        "WAYBILL": "waybill_number",

        # shipment_date
        "TARİH": "shipment_date",
        "TARIH": "shipment_date",
        "SEVKİYAT TARİHİ": "shipment_date",
        "SEVKIYAT TARIHI": "shipment_date",
        "SHIPMENT_DATE": "shipment_date",
        "DATE": "shipment_date",

        # sender
        "GÖNDERİCİ FİRMA/ŞAHIS": "sender",
        "GONDERICI FIRMA/SAHIS": "sender",
        "GÖNDERİCİ FİRMA / ŞAHIS": "sender",
        "GONDERICI FIRMA / SAHIS": "sender",
        "GÖNDERİCİ": "sender",
        "GONDERICI": "sender",
        "SENDER": "sender",

        # destination
        "ÜLKE-VARIŞ NOKTASI": "destination",
        "ULKE-VARIS NOKTASI": "destination",
        "ÜLKE - VARIŞ NOKTASI": "destination",
        "ULKE - VARIS NOKTASI": "destination",
        "ÜLKE/VARIŞ NOKTASI": "destination",
        "ULKE/VARIS NOKTASI": "destination",
        "ÜLKE": "destination",
        "ULKE": "destination",
        "VARIŞ NOKTASI": "destination",
        "VARIS NOKTASI": "destination",
        "DESTINATION": "destination",
        "COUNTRY": "destination",

        # piece_count
        "PARÇA": "piece_count",
        "PARCA": "piece_count",
        "PARÇA SAYISI": "piece_count",
        "PARCA SAYISI": "piece_count",
        "PIECE": "piece_count",
        "PIECES": "piece_count",
        "PIECE_COUNT": "piece_count",

        # collected_by
        "TOPLAYAN": "collected_by",
        "TOPLAYAN KURYE": "collected_by",
        "KURYE": "collected_by",
        "COLLECTED_BY": "collected_by",

        # delivered
        "TESLİM EDİLDİ": "delivered",
        "TESLIM EDILDI": "delivered",
        "TESLİM": "delivered",
        "TESLIM": "delivered",
        "TESLİM DURUMU": "delivered",
        "TESLIM DURUMU": "delivered",
        "DELIVERED": "delivered",

        # receiver
        "ALICI FIRMA/ŞAHIS": "receiver",
        "ALICI FIRMA/SAHIS": "receiver",
        "ALICI FİRMA/ŞAHIS": "receiver",
        "ALICI FİRMA / ŞAHIS": "receiver",
        "ALICI": "receiver",
        "RECEIVER": "receiver",

        # euro_amount
        "EURO": "euro_amount",
        "EURO TUTARI": "euro_amount",
        "EURO_AMOUNT": "euro_amount",
        "TUTAR": "euro_amount",

        # exchange_rate
        "KUR": "exchange_rate",
        "DÖVİZ KURU": "exchange_rate",
        "DOVIZ KURU": "exchange_rate",
        "EXCHANGE_RATE": "exchange_rate",
        "RATE": "exchange_rate",
    }

    def _normalize_header(self, header):
        """Sütun başlığını temizleyip standart büyük harfe çevirir."""
        if not header:
            return ""
        return str(header).strip().upper()

    def _normalize_text_field(self, raw_value, fallback="-"):
        if pd.isna(raw_value):
            return fallback
        value_str = str(raw_value).strip()
        if value_str.upper() in self.EMPTY_PLACEHOLDER_VALUES:
            return fallback
        return value_str

    def _normalize_waybill_number(self, raw_value):
        if pd.isna(raw_value):
            return None
        if isinstance(raw_value, float):
            try:
                raw_value = int(raw_value)
            except (ValueError, OverflowError):
                pass
        value_str = str(raw_value).strip()
        return value_str if value_str else None

    def _parse_date(self, raw_value):
        if pd.isna(raw_value):
            return PLACEHOLDER_DATE
        try:
            return pd.to_datetime(raw_value).date()
        except (ValueError, TypeError):
            return PLACEHOLDER_DATE

    def _parse_int(self, raw_value):
        if pd.isna(raw_value):
            return None
        try:
            val_str = str(raw_value).strip().replace(" ", "")
            val = int(float(val_str))
            return val if val >= 0 else None
        except (ValueError, TypeError, OverflowError):
            return None

    def _parse_decimal(self, raw_value, decimal_places=2):
        if pd.isna(raw_value):
            return None
        try:
            val_str = str(raw_value).strip()
            for token in ["€", "$", "TL", "tl", "EUR", "USD", " "]:
                val_str = val_str.replace(token, "")
            # Türkçe virgüllü format "12,50" -> "12.50"
            if "," in val_str and "." in val_str:
                val_str = val_str.replace(".", "").replace(",", ".")
            elif "," in val_str:
                val_str = val_str.replace(",", ".")
            d = Decimal(val_str)
            if d < 0:
                return None
            return round(d, decimal_places)
        except (InvalidOperation, ValueError, TypeError):
            return None

    def _parse_boolean(self, raw_value):
        if pd.isna(raw_value):
            return False
        val_str = str(raw_value).strip().lower()
        if val_str in {"1", "true", "t", "evet", "e", "yes", "y", "teslim", "teslim edildi", "ok", "+"}:
            return True
        return False

    def post(self, request, *args, **kwargs):
        upload_serializer = WaybillExcelUploadSerializer(data=request.data)
        upload_serializer.is_valid(raise_exception=True)
        excel_file = upload_serializer.validated_data["file"]

        try:
            file_name = excel_file.name.lower()
            if file_name.endswith(".csv"):
                df = pd.read_csv(excel_file)
            elif file_name.endswith(".xls"):
                df = pd.read_excel(excel_file, engine="xlrd")
            else:
                df = pd.read_excel(excel_file, engine="openpyxl")
        except Exception as e:
            return Response(
                {"detail": f"Dosya okunamadı: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if df.empty:
            return Response(
                {"detail": "Excel dosyası boş."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Excel sütunlarını model alanlarına eşle
        field_to_col = {}
        for col in df.columns:
            norm_col = self._normalize_header(col)
            if norm_col in self.COLUMN_ALIASES:
                target_field = self.COLUMN_ALIASES[norm_col]
                if target_field not in field_to_col:
                    field_to_col[target_field] = col

        if "waybill_number" not in field_to_col:
            return Response(
                {
                    "detail": "Excel dosyasında AWB / Konşimento No sütunu bulunamadı.",
                    "available_columns": list(df.columns),
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        to_update = []
        to_create = []
        errors = []

        existing_numbers = set(
            Waybill.objects.values_list("waybill_number", flat=True)
        )

        for index, row in df.iterrows():
            excel_row_number = index + 2

            try:
                awb_col = field_to_col["waybill_number"]
                waybill_number = self._normalize_waybill_number(row[awb_col])
                if not waybill_number:
                    raise ValueError("AWB / Konşimento numarası boş olamaz.")

                # Opsiyonel alanları çek ve normalize et
                shipment_date = (
                    self._parse_date(row[field_to_col["shipment_date"]])
                    if "shipment_date" in field_to_col
                    else PLACEHOLDER_DATE
                )

                sender = (
                    self._normalize_text_field(row[field_to_col["sender"]])
                    if "sender" in field_to_col
                    else "-"
                )

                destination = (
                    self._normalize_text_field(row[field_to_col["destination"]])
                    if "destination" in field_to_col
                    else "-"
                )

                piece_count = (
                    self._parse_int(row[field_to_col["piece_count"]])
                    if "piece_count" in field_to_col
                    else None
                )

                collected_by = (
                    self._normalize_text_field(row[field_to_col["collected_by"]])
                    if "collected_by" in field_to_col
                    else "-"
                )

                delivered = (
                    self._parse_boolean(row[field_to_col["delivered"]])
                    if "delivered" in field_to_col
                    else False
                )

                receiver = (
                    self._normalize_text_field(row[field_to_col["receiver"]])
                    if "receiver" in field_to_col
                    else "-"
                )

                euro_amount = (
                    self._parse_decimal(row[field_to_col["euro_amount"]], decimal_places=2)
                    if "euro_amount" in field_to_col
                    else None
                )

                exchange_rate = (
                    self._parse_decimal(row[field_to_col["exchange_rate"]], decimal_places=4)
                    if "exchange_rate" in field_to_col
                    else None
                )

                data = {
                    "waybill_number": waybill_number,
                    "shipment_date": shipment_date,
                    "sender": sender,
                    "destination": destination,
                    "piece_count": piece_count,
                    "collected_by": collected_by,
                    "delivered": delivered,
                    "receiver": receiver,
                    "euro_amount": euro_amount,
                    "exchange_rate": exchange_rate,
                }

                if waybill_number in existing_numbers:
                    to_update.append(data)
                else:
                    to_create.append(Waybill(**data))
                    existing_numbers.add(waybill_number)

            except (ValueError, KeyError, TypeError) as e:
                errors.append({"row": excel_row_number, "error": str(e)})

        created_count = 0
        updated_count = 0

        try:
            with transaction.atomic():
                if to_create:
                    Waybill.objects.bulk_create(to_create, batch_size=500)
                    created_count = len(to_create)

                for data in to_update:
                    _, created = Waybill.objects.update_or_create(
                        waybill_number=data["waybill_number"],
                        defaults=data,
                    )
                    updated_count += 1

        except Exception as e:
            return Response(
                {"detail": f"Veritabanı işlemi sırasında hata oluştu: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        return Response(
            {
                "detail": "Yükleme tamamlandı.",
                "created": created_count,
                "updated": updated_count,
                "error_count": len(errors),
                "errors": errors,
            },
            status=status.HTTP_200_OK,
        )


# --------------------------------------------------------------------------
# 2) LİSTELEME VE FİLTRELEME ENDPOINT'İ
# --------------------------------------------------------------------------

class WaybillListView(generics.ListAPIView):
    """
    GET /api/waybills/?start_date=...&end_date=...&delivered=true&page=1

    Tarih aralığına, teslim durumuna ve eksik-veri durumuna göre filtreleme + sayfalama.
    """
    serializer_class = WaybillSerializer
    pagination_class = WaybillPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["waybill_number", "sender", "receiver", "destination", "collected_by"]
    ordering_fields = [
        "waybill_number",
        "shipment_date",
        "sender",
        "destination",
        "piece_count",
        "collected_by",
        "delivered",
        "receiver",
        "euro_amount",
        "exchange_rate",
    ]
    ordering = ["-shipment_date", "-id"]

    def get_queryset(self):
        return filter_waybills_queryset(self.request)

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        summary = queryset.aggregate(
            total_count=Count("id"),
            total_pieces=Sum("piece_count"),
            total_euro=Sum("euro_amount"),
            delivered_count=Count("id", filter=Q(delivered=True)),
        )

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            response = self.get_paginated_response(serializer.data)
            response.data["summary"] = {
                "total_count": summary["total_count"] or 0,
                "total_pieces": summary["total_pieces"] or 0,
                "total_euro": float(summary["total_euro"] or 0),
                "delivered_count": summary["delivered_count"] or 0,
            }
            return response

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


# --------------------------------------------------------------------------
# 3) FİLTRELENMİŞ SONUÇLARI EXCEL OLARAK DIŞA AKTARMA
# --------------------------------------------------------------------------

class WaybillExportView(APIView):
    """
    GET /api/waybills/export/?start_date=...&end_date=...&delivered=...
    """

    def get(self, request, *args, **kwargs):
        queryset = filter_waybills_queryset(request)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Konşimentolar"

        headers = [
            "Tarih",
            "AWB",
            "Gönderici Firma/Şahıs",
            "Ülke-Varış Noktası",
            "Parça",
            "Toplayan",
            "Teslim Edildi",
            "Alıcı Firma/Şahıs",
            "Euro",
            "Kur",
            "Ödenen Tutar (TL)",
        ]
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill

        for waybill in queryset.iterator():
            shipment_date_str = (
                waybill.shipment_date.strftime("%d.%m.%Y")
                if waybill.shipment_date != PLACEHOLDER_DATE
                else "VERİ YOK"
            )
            ws.append([
                shipment_date_str,
                waybill.waybill_number,
                waybill.sender,
                waybill.destination,
                waybill.piece_count if waybill.piece_count is not None else "-",
                waybill.collected_by,
                "Evet" if waybill.delivered else "Hayır",
                waybill.receiver,
                float(waybill.euro_amount) if waybill.euro_amount is not None else "-",
                float(waybill.exchange_rate) if waybill.exchange_rate is not None else "-",
                waybill.payment_amount if waybill.payment_amount is not None else "-",
            ])

        column_widths = [15, 18, 28, 22, 10, 18, 14, 28, 12, 10, 16]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        start_date_param = request.query_params.get("start_date")
        end_date_param = request.query_params.get("end_date")
        filename = f"konsimentolar_{start_date_param or 'tumu'}_{end_date_param or 'tumu'}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response


# --------------------------------------------------------------------------
# 4) TEKİL KAYIT: GÖRÜNTÜLEME, GÜNCELLEME, SİLME
# --------------------------------------------------------------------------

class WaybillDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    GET    /api/waybills/<id>/
    PATCH  /api/waybills/<id>/
    PUT    /api/waybills/<id>/
    DELETE /api/waybills/<id>/
    """
    queryset = Waybill.objects.all()
    serializer_class = WaybillSerializer